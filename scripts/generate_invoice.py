"""
Generate invoice xlsx files for the Gymshark Phone Box proposal.
Producer: AGV Miami (legal entity AGV Miami, LLC).

Outputs (run: python3 scripts/generate_invoice.py):
  1. public/invoices/GS-PHONEBOX-001-V3.0.xlsx — Single-tab V3.0.
  2. public/invoices/GS-PHONEBOX-001-V3.1.xlsx — Single-tab V3.1.
  3. public/invoices/GS-PHONEBOX-001-V4.0.xlsx — Three tabs: V4.0
     (current), V3.0, V2.0.
  4. public/invoices/GS-PHONEBOX-001-VERSION-HISTORY.xlsx — Five
     tabs: V4.0, V3.1, V3.0, V2.0, V1.0.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_V3 = "public/invoices/GS-PHONEBOX-001-V3.0.xlsx"
OUTPUT_V31 = "public/invoices/GS-PHONEBOX-001-V3.1.xlsx"
OUTPUT_V40 = "public/invoices/GS-PHONEBOX-001-V4.0.xlsx"
OUTPUT_HISTORY = "public/invoices/GS-PHONEBOX-001-VERSION-HISTORY.xlsx"

GOLD = "D4AF37"
DARK = "0C0C10"
INK = "1A1A22"
MUTED = "8A8A95"
WHITE = "FFFFFF"

f_eyebrow = Font(name="Helvetica", size=9, bold=True, color=GOLD)
f_label = Font(name="Helvetica", size=9, bold=True, color=MUTED)
f_meta = Font(name="Helvetica", size=11, color=DARK)
f_phase = Font(name="Helvetica", size=11, bold=True, color=WHITE)
f_line_name = Font(name="Helvetica", size=10, bold=True, color=DARK)
f_line_desc = Font(name="Helvetica", size=9, color=MUTED)
f_amount = Font(name="Helvetica", size=10, bold=True, color=DARK)
f_subtotal = Font(name="Helvetica", size=10, bold=True, color=GOLD)
f_credit = Font(name="Helvetica", size=10, bold=True, color="C44A4A")
f_footer = Font(name="Helvetica", size=8, color=MUTED, italic=True)

fill_phase = PatternFill("solid", start_color=INK, end_color=INK)
fill_subtotal = PatternFill("solid", start_color="F5F1E0", end_color="F5F1E0")
fill_total = PatternFill("solid", start_color=GOLD, end_color=GOLD)

TAB_GOLD = "D4AF37"
TAB_YELLOW = "F4C430"
TAB_GREEN = "6FCF97"
TAB_TEAL = "00C2A8"
TAB_MUTED = "8A8A95"


def setup(ws):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 4


def header_block(ws, row, version, issue_date, scope_summary, activation_summary):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="AGV MIAMI")
    cell.font = Font(name="Helvetica", size=24, bold=True, color=DARK)
    ws.row_dimensions[row].height = 36
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=f"EXPERIENTIAL FABRICATION & PRODUCTION  ·  PROPOSAL INVOICE  ·  {version}")
    cell.font = Font(name="Helvetica", size=9, bold=True, color=GOLD)
    row += 2

    ws.cell(row=row, column=2, value="PRODUCER").font = f_label
    ws.cell(row=row, column=3, value="CLIENT").font = f_label
    ws.cell(row=row, column=4, value="PROJECT").font = f_label
    row += 1

    ws.cell(row=row, column=2, value="AGV Miami, LLC").font = f_meta
    ws.cell(row=row, column=3, value="Gymshark Ltd.").font = f_meta
    ws.cell(row=row, column=4, value="GS-PHONEBOX-001").font = f_meta
    row += 1

    ws.cell(row=row, column=2, value="1440 Church St").font = f_meta
    ws.cell(row=row, column=3, value="Attn: Authorized Billing Contact").font = f_meta
    ws.cell(row=row, column=4, value=version).font = f_meta
    row += 1

    ws.cell(row=row, column=2, value="Bohemia, NY 11716").font = f_meta
    ws.cell(row=row, column=3, value="c/o Ominto Studio").font = f_meta
    ws.cell(row=row, column=4, value=f"Issue Date: {issue_date}").font = f_meta
    row += 1

    ws.cell(row=row, column=2, value="jclarkson@agvmiami.com").font = f_meta
    ws.cell(row=row, column=4, value=activation_summary[0]).font = f_meta
    row += 1

    ws.cell(row=row, column=2, value="info@agoravisuals.com").font = f_meta
    ws.cell(row=row, column=4, value=activation_summary[1] if len(activation_summary) > 1 else "").font = f_meta
    row += 2

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=f"{version.split()[1]} SCOPE")
    cell.font = f_eyebrow
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=scope_summary)
    cell.font = Font(name="Helvetica", size=10, color=DARK)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = max(40, 14 * (1 + len(scope_summary) // 90))
    row += 2
    return row


def phase_header(ws, row, label):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value=label)
    c.font = f_phase
    c.fill = fill_phase
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24
    return row + 1


def line_item(ws, row, name, desc, amount):
    ws.cell(row=row, column=2, value=name).font = f_line_name
    ws.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(row=row, column=3, value=desc).font = f_line_desc
    ws.cell(row=row, column=3).alignment = Alignment(vertical="top", wrap_text=True)
    c = ws.cell(row=row, column=4, value=amount)
    c.font = f_amount
    c.number_format = '"$"#,##0'
    c.alignment = Alignment(horizontal="right", vertical="top")
    ws.row_dimensions[row].height = 36
    return row + 1


def subtotal_row(ws, row, label, amount):
    for col in (2, 3, 4):
        ws.cell(row=row, column=col).fill = fill_subtotal
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value=label)
    c.font = f_subtotal
    c.alignment = Alignment(horizontal="right", indent=1)
    c2 = ws.cell(row=row, column=4, value=amount)
    c2.font = f_subtotal
    c2.number_format = '"$"#,##0'
    c2.alignment = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 22
    return row + 1


def gross_total_row(ws, row, label, amount):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value=label)
    c.font = Font(name="Helvetica", size=12, bold=True, color=DARK)
    c.alignment = Alignment(horizontal="right", indent=1)
    c2 = ws.cell(row=row, column=4, value=amount)
    c2.font = Font(name="Helvetica", size=12, bold=True, color=DARK)
    c2.number_format = '"$"#,##0'
    c2.alignment = Alignment(horizontal="right")
    return row + 1


def credit_row(ws, row, label, amount):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value=label)
    c.font = f_credit
    c.alignment = Alignment(horizontal="right", indent=1)
    c2 = ws.cell(row=row, column=4, value=amount)
    c2.font = f_credit
    c2.number_format = '"-$"#,##0'
    c2.alignment = Alignment(horizontal="right")
    return row + 1


def net_total_bar(ws, row, label, amount):
    for col in (2, 3, 4):
        ws.cell(row=row, column=col).fill = fill_total
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c = ws.cell(row=row, column=2, value=label)
    c.font = Font(name="Helvetica", size=14, bold=True, color=DARK)
    c.alignment = Alignment(horizontal="right", indent=1, vertical="center")
    c2 = ws.cell(row=row, column=4, value=amount)
    c2.font = Font(name="Helvetica", size=18, bold=True, color=DARK)
    c2.number_format = '"$"#,##0'
    c2.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 32
    return row + 1


def payment_block(ws, row, deposit, balance, balance_due_date, version_short):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value="PAYMENT TERMS")
    cell.font = f_eyebrow
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    body = (
        f"60% deposit (${deposit:,}) due upon Client's written approval of this Scope of Work (Proposal execution). "
        f"40% balance (${balance:,}) due five (5) business days prior to first activation installation "
        f"({balance_due_date}). "
        f"Payment exclusively via ACH electronic transfer or domestic wire transfer; ACH/wire details issued directly to the Client billing contact upon SoW approval. "
        f"Reference GS-PHONEBOX-001 {version_short} on all remittances."
    )
    cell = ws.cell(row=row, column=2, value=body)
    cell.font = Font(name="Helvetica", size=10, color=DARK)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 70
    return row + 2


def footer(ws, row, version_short):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=2, value=(
        f"AGV Miami, LLC  ·  1440 Church St, Bohemia, NY 11716  ·  jclarkson@agvmiami.com  ·  This invoice is issued in connection with proposal "
        f"GS-PHONEBOX-001 {version_short} and incorporates by reference the executed Master Services Agreement between AGV Miami, LLC and Gymshark Ltd."
    ))
    cell.font = f_footer
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.row_dimensions[row].height = 32


# ---------- V3.0 SHEET (current) ----------
def build_v3_sheet(ws):
    setup(ws)
    ws.sheet_properties.tabColor = TAB_GOLD
    row = header_block(
        ws, row=1,
        version="Version 3.0",
        issue_date="April 29, 2026",
        scope_summary=(
            "Single-event programme: fully-painted pink British phone box with interactive call/voucher/photo tech, "
            "deployed for a one-day street activation at Lincoln Road Mall, Miami Beach on Friday, "
            "July 17, 2026 (Saturday, July 18 held as a weather contingency). Light touchup post-Miami; logistical "
            "return to NYC handoff destination — no NYC retail activation, no respray/rewrap package."
        ),
        activation_summary=("Activation: Friday, July 17, 2026", "Venue: Lincoln Road Mall"),
    )

    row = phase_header(ws, row, "Phase 01  —  The Phone Box (Painted Finish)")
    row = line_item(ws, row, "Phone Box Structural Shell", "Custom scenic fabrication, marine-grade paint in Gymshark Pink, two-piece modular construction", 14500)
    row = line_item(ws, row, "4-Sided Illuminated Lightbox Signage", "LED-backlit translucent face panels on all four sides; weatherproof housing", 6800)
    row = line_item(ws, row, "Glass & Semi-Transparent Vinyl Panels", "Tempered glass on front door + sides with custom-printed semi-transparent privacy vinyl", 2900)
    row = line_item(ws, row, "Painted Pink Interior (V3 — paint, not vinyl)", "Four-wall matte paint finish (Gymshark Pink), dimensional bum mirror with scripted messaging. Replaces V2.0 vinyl lining for durability and finish quality.", 3800)
    row = line_item(ws, row, "Concealed Sliding Prize Door", "Motorised, colour-matched, no visible handle on public face", 3400)
    row = line_item(ws, row, "Interior Finishes", "Aluminium chequer-plate flooring, dome light, guest seat, branded analogue phone", 2750)
    row = line_item(ws, row, "Engineering, Structural Calcs & Shop Drawings", "Wind-load, ballast plan, electrical schematics, CAD shop drawings, venue-compliance package", 4000)
    row = subtotal_row(ws, row, "Subtotal — The Phone Box", 38150)

    row = phase_header(ws, row, "Phase 02  —  Interactive Tech + Content Capture")
    row = line_item(ws, row, "Pre-Recorded Call-Response (IVR) System", "Multi-branch scripting, licensed voice talent, keypad mapping, redundant win/lose logic, QA", 5500)
    row = line_item(ws, row, "Live Call / Walkie-Talkie Relay", "Dual-mode encrypted relay, backup uplink, on-site producer headset, external broadcast", 2900)
    row = line_item(ws, row, "Ceiling Dual Camera System", "Two 4K video+audio cameras, cloud storage, live-preview monitoring, disclosure signage", 4500)
    row = line_item(ws, row, "Selfie & Belfie Mobile Mounts", "Two pink wall-mounted holders with motion-triggered ring-light inserts, dual-height", 1800)
    row = line_item(ws, row, "Thermal Voucher Printer & Shelf Mount", "Dual-speed thermal ticket printer, pink wrap, redundant roll inventory, firmware", 2800)
    row = line_item(ws, row, "Content Capture & Media Handoff Platform", "Cloud workspace, colour-corrected proxies, organised file-naming, 48-hour handoff", 2500)
    row = subtotal_row(ws, row, "Subtotal — Interactive Tech + Content Capture", 20000)

    row = phase_header(ws, row, "Phase 03  —  Branding, Signage & Consumables")
    row = line_item(ws, row, "Disclosure & Wayfinding Signage", "Camera-disclosure sign, queue management decals, brand lockup callouts in brand palette", 950)
    row = line_item(ws, row, "Venue Dressing Kit", "Portable A-frame signage, branded stanchions, pavement decals for the Lincoln Road Mall street activation", 1750)
    row = line_item(ws, row, "Daily Consumables & Spares Kit", "Voucher paper rolls, cleaning supplies, sanitisation wipes, touch-up paint, vinyl repair, pink gaffer", 2500)
    row = subtotal_row(ws, row, "Subtotal — Branding, Signage & Consumables", 5200)

    row = phase_header(ws, row, "Phase 04a  —  The Activation (Lincoln Road Mall, Miami Beach)")
    row = line_item(ws, row, "Inbound Logistics & Install", "Truck, rigging hardware, 3-person install crew, permit-window supervisor, 4–6 hr install window", 5800)
    row = line_item(ws, row, "On-Site Technicians", "Dedicated lead technician plus rotating second tech for rush windows and breaks", 2800)
    row = line_item(ws, row, "Same-Day Strike & Outbound Freight", "Complete de-installation, module breakdown, crated outbound freight, full site restoration", 3300)
    row = subtotal_row(ws, row, "Subtotal — The Activation (Lincoln Road Mall)", 11900)

    row = phase_header(ws, row, "Phase 04b  —  Logistics to NYC (Asset Transfer Only)")
    row = line_item(ws, row, "Climate-Controlled Warehouse Hold", "Secure climate-controlled storage at the AGV Miami NYC staging facility between Miami strike and NYC handoff", 2400)
    row = line_item(ws, row, "Inter-City Freight (Climate-Controlled Truck)", "Miami → AGV Miami NYC staging facility with real-time GPS tracking, two-driver rotation, handoff documentation", 5500)
    row = line_item(ws, row, "Light Touchup After First Activation (V3 — new line)", "Single light cosmetic touchup pass after Miami activation: scuff/scratch repair, paint colour-match, lightbox edge cleanup, IVR/camera/printer functional re-test", 3000)
    row = line_item(ws, row, "Final Delivery to NYC Handoff Destination", "Climate-controlled final-mile delivery from AGV Miami NYC staging to client's designated NYC handoff address. Asset transfer only — no installation, commissioning, or activation.", 2000)
    row = subtotal_row(ws, row, "Subtotal — Logistics to NYC", 12900)

    row = phase_header(ws, row, "Phase 05  —  Project Management & Client Services")
    row = line_item(ws, row, "Project Management Fee", "Dedicated senior producer, weekly status reporting, milestone tracking, Change Order administration, vendor / venue liaison, COI + insurance coordination, post-event reconciliation", 12000)
    row = subtotal_row(ws, row, "Subtotal — Project Management", 12000)

    row += 1
    row = gross_total_row(ws, row, "Gross Production Investment", 100150)
    row = credit_row(ws, row, "Less: Preferred Partner Credit", -5000)
    row = net_total_bar(ws, row, "NET V3.0 PRODUCTION INVESTMENT", 95150)
    row += 1

    row = payment_block(ws, row, deposit=57090, balance=38060, balance_due_date="on or before Friday, July 10, 2026", version_short="V3.0")
    footer(ws, row, "V3.0")


# ---------- V3.1 SHEET (proposed update — yellow palette + scope deltas) ----------
def build_v31_sheet(ws):
    setup(ws)
    ws.sheet_properties.tabColor = TAB_YELLOW
    row = header_block(
        ws, row=1,
        version="Version 3.1",
        issue_date="May 7, 2026",
        scope_summary=(
            "V3.1 proposed update — yellow Miami palette throughout (paint, vinyl, lightbox, signage), "
            "operating hours extended to 11 AM–8 PM, back-of-house staff compartment with separate back "
            "access door, simplified consumer walkie-talkie pair, basic selfie/belfie mounts (no ring lights), "
            "Content Capture & Media Handoff Platform removed, 200-unit custom packaging added, and a local "
            "NYC install crew + in-store fixture setup + store wayfinding for the post-Miami delivery. "
            "Activation remains Friday, July 17, 2026 at Lincoln Road Mall (Saturday, July 18 "
            "held as a weather contingency)."
        ),
        activation_summary=("Activation: Friday, July 17, 2026, 11 AM–8 PM", "Venue: Lincoln Road Mall"),
    )

    row = phase_header(ws, row, "Phase 01  —  The Phone Box (Painted Yellow Finish)")
    row = line_item(ws, row, "Phone Box Structural Shell", "Custom scenic fabrication, marine-grade paint in Miami yellow (PMS TBC by Ominto), two-piece modular construction", 14500)
    row = line_item(ws, row, "4-Sided Illuminated Lightbox Signage", "LED-backlit translucent face panels on all four sides; weatherproof housing", 6800)
    row = line_item(ws, row, "Glass & Semi-Transparent Vinyl Panels", "Tempered glass on front door + sides with custom-printed yellow-palette privacy vinyl", 2900)
    row = line_item(ws, row, "Painted Yellow Interior (V3.1 — yellow palette)", "Four-wall matte paint finish in Miami yellow, dimensional bum mirror with scripted messaging. Yellow PMS callout to be confirmed by Ominto.", 3800)
    row = line_item(ws, row, "Concealed Sliding Prize Door", "Motorised, colour-matched, no visible handle on public face — for slot-style product handoff from the back-of-house staff compartment", 3400)
    row = line_item(ws, row, "Interior Finishes", "Aluminium chequer-plate flooring, dome light, guest seat, branded analogue phone", 2750)
    row = line_item(ws, row, "Engineering, Structural Calcs & Shop Drawings", "Wind-load, ballast plan, electrical schematics, CAD shop drawings, venue-compliance package", 4000)
    row = line_item(ws, row, "Back-of-House Compartment + Back Access Door (V3.1 — new)", "Interior partition wall + hinged back-access door + small staff stoop, allowing one Gymshark-supplied team member to occupy a private back compartment and dispense product through the motorised slot to the guest. Talent supplied by Client per the proposal exclusions.", 2400)
    row = subtotal_row(ws, row, "Subtotal — The Phone Box", 40550)

    row = phase_header(ws, row, "Phase 02  —  Interactive Tech (V3.1 — simplified)")
    row = line_item(ws, row, "Pre-Recorded Call-Response (IVR) System", "Multi-branch scripting, licensed voice talent, keypad mapping, redundant win/lose logic, QA", 5500)
    row = line_item(ws, row, "Walkie-Talkie Pair (V3.1 — consumer-grade)", "Off-the-shelf Motorola pair: one inside the booth for the guest, one with the on-site Gymshark athlete. Includes spares. Replaces the V3.0 dual-mode encrypted relay with backup uplink.", 400)
    row = line_item(ws, row, "Ceiling Dual Camera System", "Two 4K video+audio cameras, cloud storage, live-preview monitoring, disclosure signage", 4500)
    row = line_item(ws, row, "Selfie & Belfie Mobile Mounts (V3.1 — basic)", "Two simple wall-mounted phone holders for guests' own phones. Painted to match interior. Replaces the V3.0 motion-triggered ring-light variant.", 600)
    row = line_item(ws, row, "Thermal Voucher Printer & Shelf Mount", "Dual-speed thermal ticket printer, wrapped, redundant roll inventory, firmware", 2800)
    row = line_item(ws, row, "(Removed in V3.1) Content Capture & Media Handoff Platform", "Cloud workspace + colour-corrected proxies + 48-hour delivery — REMOVED at Client request; Client supplies own media pipeline. Raw camera files handed over on drive at strike at no extra cost.", 0)
    row = subtotal_row(ws, row, "Subtotal — Interactive Tech", 13800)

    row = phase_header(ws, row, "Phase 03  —  Branding, Signage & Consumables (V3.1 — packaging added)")
    row = line_item(ws, row, "Disclosure & Wayfinding Signage", "Camera-disclosure sign, queue management decals, brand lockup callouts in yellow palette", 950)
    row = line_item(ws, row, "Venue Dressing Kit", "Portable A-frame signage, branded stanchions, pavement decals for the Lincoln Road Mall street activation, yellow palette", 1750)
    row = line_item(ws, row, "Daily Consumables & Spares Kit", "Voucher paper rolls, cleaning supplies, sanitisation wipes, yellow touch-up paint, vinyl repair, gaffer", 2500)
    row = line_item(ws, row, "Custom Product Packaging (V3.1 — new) — 200 units", "Yellow-palette product boxes, structural board, full-colour offset print, 300 × 200 × 70 mm, flat-packed for load-in. Per-unit scales linearly above 100; under 100 the per-unit cost climbs.", 4800)
    row = subtotal_row(ws, row, "Subtotal — Branding, Signage & Consumables", 10000)

    row = phase_header(ws, row, "Phase 04a  —  The Activation (Lincoln Road Mall, 11 AM–8 PM, V3.1)")
    row = line_item(ws, row, "Inbound Logistics & Install", "Truck, rigging hardware, 3-person install crew, permit-window supervisor, 4–6 hr install window", 5800)
    row = line_item(ws, row, "On-Site Technicians (V3.1 — extended to 8 PM)", "Dedicated lead technician plus rotating second tech for rush windows and breaks, covering an extended 11 AM–8 PM operating window. Includes pre-open soundcheck and post-close shutdown.", 3000)
    row = line_item(ws, row, "Same-Day Strike & Outbound Freight", "Complete de-installation, module breakdown, crated outbound freight within the Lincoln Road BID's contracted strike window. Site walk-through with the BID operations team.", 3300)
    row = subtotal_row(ws, row, "Subtotal — The Activation (Lincoln Road Mall)", 12100)

    row = phase_header(ws, row, "Phase 04b  —  Logistics + Local Install at NYC Retail (V3.1)")
    row = line_item(ws, row, "Climate-Controlled Warehouse Hold", "Secure climate-controlled storage at the AGV Miami NYC staging facility between Miami strike and NYC delivery", 2400)
    row = line_item(ws, row, "Inter-City Freight (Climate-Controlled Truck)", "Miami → AGV Miami NYC staging facility with real-time GPS tracking, two-driver rotation, handoff documentation", 5500)
    row = line_item(ws, row, "Light Touchup After First Activation", "Single light cosmetic touchup pass after the Miami activation: scuff/scratch repair, paint colour-match, lightbox edge cleanup, IVR/camera/printer functional re-test. Brief refurbish, not a full repaint.", 3000)
    row = line_item(ws, row, "White-Glove Delivery to NYC Retail Location", "Climate-controlled final-mile delivery from AGV Miami NYC staging to the Client's nominated NYC retail address (Bond St assumed unless otherwise specified). Scheduled overnight or pre-open per store operations.", 2000)
    row = line_item(ws, row, "Retail-Environment Install Crew (V3.1 — new)", "2-person retail-environment install crew with floor-protection, retail-grade hand-tool kit, store operations liaison. 4-hour install window scheduled to avoid trading hours.", 2300)
    row = line_item(ws, row, "In-Store Fixture Setup (V3.1 — new)", "Anchoring to retail-spec floor (no permanent penetration), electrical drop, IVR + camera + printer + lightbox bring-up, on-floor commissioning sign-off with the store manager.", 2100)
    row = line_item(ws, row, "Store Wayfinding & Footfall Driver Kit (V3.1 — new)", "Sidewalk A-frame outside the store entrance, branded window-vinyl tease pointing to the Phone Box installation inside, in-store directional decals, and a printed/social CTA pack to drive walk-in footfall during the launch window.", 1000)
    row = subtotal_row(ws, row, "Subtotal — Logistics + NYC Retail Install", 18300)

    row = phase_header(ws, row, "Phase 05  —  Project Management & Client Services")
    row = line_item(ws, row, "Project Management Fee", "Dedicated senior producer, weekly status reporting, milestone tracking, Change Order administration, vendor / venue liaison, COI + insurance coordination, post-event reconciliation", 12000)
    row = subtotal_row(ws, row, "Subtotal — Project Management", 12000)

    row += 1
    row = gross_total_row(ws, row, "Gross Production Investment", 106750)
    row = credit_row(ws, row, "Less: Preferred Partner Credit", -5000)
    row = net_total_bar(ws, row, "NET V3.1 PRODUCTION INVESTMENT", 101750)
    row += 1

    row = payment_block(ws, row, deposit=61050, balance=40700, balance_due_date="on or before Friday, July 10, 2026", version_short="V3.1")
    footer(ws, row, "V3.1 — proposed update, awaiting Client sign-off")


# ---------- V4.0 SHEET (current proposed — V3.1 clarifications + deeper VE) ----------
def build_v40_sheet(ws):
    setup(ws)
    ws.sheet_properties.tabColor = TAB_GREEN
    row = header_block(
        ws, row=1,
        version="Version 4.0",
        issue_date="May 8, 2026",
        scope_summary=(
            "Single-event programme: a fully-painted yellow British phone box equipped with interactive "
            "photo, voucher, and call-response technology, deployed for a one-day street activation at "
            "Lincoln Road Mall, Miami Beach on Friday, July 17, 2026, 11 AM–8 PM (Saturday, July 18 held "
            "as a weather contingency). After Miami the booth ships to AGV Miami's NY shop for a light "
            "touchup and is white-glove delivered to the Gymshark NYC flagship store by Thursday, "
            "July 23, 2026. Booth dimensions per Ominto V2 design: 164 × 94 cm body + 20 cm lightbox "
            "header. Hinged back distribution door for in-booth staff to dispense product directly. "
            "$5,000 Preferred Partner Credit applied. Local NYC install team available as an optional "
            "add-on."
        ),
        activation_summary=("Activation: Friday, July 17, 2026, 11 AM–8 PM", "Venue: Lincoln Road Mall"),
    )

    row = phase_header(ws, row, "Phase 01  —  The Phone Box")
    row = line_item(ws, row, "Phone Box Structural Shell", "Custom scenic fabrication per Ominto V2 design pack. Footprint 164 × 94 cm (~5'4\" × 3'1\"); body height 213 cm (~7'0\"); lightbox header +20 cm (total ≈7'8\"). Two-piece modular construction (staff compartment 70 cm deep + guest compartment 94 cm deep). Marine-grade paint in Miami yellow per Ominto brand guidelines.", 14500)
    row = line_item(ws, row, "4-Sided Illuminated Lightbox Signage", "20 cm lightbox header on all four faces. LED-backlit translucent face panels with weatherproof housing, dimmable driver, single-cord power feed.", 6800)
    row = line_item(ws, row, "Glass & Semi-Transparent Vinyl Panels", "Tempered glass on front door + sides per Ominto deck spec. Custom-printed semi-transparent privacy vinyl preserves the reveal moment when guests enter; finish print-matched to yellow palette.", 2900)
    row = line_item(ws, row, "Painted Yellow Interior", "Four-wall matte paint finish in Miami yellow on interior compartment. Dimensional bum mirror with scripted messaging above and below. Interior compartment dim: 65 × 86 cm (back compartment) per Ominto V2 design.", 3800)
    row = line_item(ws, row, "Hinged Back Distribution Door", "True hinged back door, brand-yellow painted, no visible handle on the public face. Allows a Gymshark-supplied staff member to operate from inside the back-of-house compartment and hand product directly to the guest.", 1800)
    row = line_item(ws, row, "Interior Finishes", "Aluminium chequer-plate (five-bar pattern) flooring, circular dome overhead light, compact guest seat (~70 cm wide), and branded analogue phone with yellow handset. Shelf below the phone receives the thermal printer.", 2750)
    row = line_item(ws, row, "Engineering, Structural Calcs & Shop Drawings", "Wind-load, ballast plan, electrical schematics, CAD shop drawings, venue-compliance package. Reconciles deck V2 dimensions (164 × 94 × 213 cm + 20 cm lightbox) at engineering lock.", 4000)
    row = line_item(ws, row, "Back-of-House Staff Compartment", "70 cm-deep × 94 cm-wide private back compartment behind the hinged distribution door. Includes interior partition wall, small staff stoop, and dispensing shelf for product handoff. Talent supplied by Client per proposal exclusions.", 1800)
    row = subtotal_row(ws, row, "Subtotal — The Phone Box", 38350)

    row = phase_header(ws, row, "Phase 02  —  Interactive Tech")
    row = line_item(ws, row, "Pre-Recorded Call-Response (IVR) System", "Multi-branch scripting, licensed voice talent, keypad mapping (* = YES, # = NO), redundant win/lose logic, QA. Wired to the analogue phone handset.", 5500)
    row = line_item(ws, row, "Two-Way Radio Pair — Motorola RMU2040", "2 × Motorola RMU2040 RM-Series 2-Way Radios (2-channel, 2-watt UHF business-grade, license-free, ~250,000 sq ft / 12-floor range, NOAA weather alerts). One inside the booth for the guest, one with the on-site Gymshark athlete; includes spare batteries.", 700)
    row = line_item(ws, row, "Ceiling Cameras — Ring Mini Indoor Plug-In", "2 × Ring Mini Indoor Security Cameras (1080p HD, two-way talk, motion detection, plug-in / no batteries, Wi-Fi, white finish). Cloud storage via Ring Protect subscription. Ceiling-mounted in front and back compartments per Ominto deck spec.", 300)
    row = line_item(ws, row, "Wall-Mounted Selfie Stations", "2 × wall-mounted selfie stations, painted to match the booth interior — face-level + belfie-angle. Each station provides a stable phone holder for guests using their own devices.", 900)
    row = line_item(ws, row, "Thermal Voucher Printer & Shelf Mount", "80mm thermal ticket printer, shelf-mounted under the analogue phone, wrapped in yellow vinyl. Voucher template designed by Gymshark; AGV Miami handles printer procurement, firmware, redundant roll inventory.", 1500)
    row = subtotal_row(ws, row, "Subtotal — Interactive Tech", 8900)

    row = phase_header(ws, row, "Phase 03  —  Branding, Signage & Consumables")
    row = line_item(ws, row, "Venue Dressing Kit", "Branded stanchion poles and ropes for crowd management at the Lincoln Road Mall street activation. Yellow palette to match booth.", 1000)
    row = line_item(ws, row, "Daily Consumables & Spares Kit", "Voucher paper rolls, cleaning supplies, sanitisation wipes, yellow touch-up paint, vinyl repair, gaffer.", 1500)
    row = line_item(ws, row, "Custom Product Packaging — 200 units", "Yellow-palette product boxes, structural board, full-colour offset print, 300 × 200 × 70 mm, flat-packed for load-in.", 4800)
    row = subtotal_row(ws, row, "Subtotal — Branding, Signage & Consumables", 7300)

    row = phase_header(ws, row, "Phase 04a  —  The Activation (Lincoln Road Mall, 11 AM–8 PM)")
    row = line_item(ws, row, "Inbound Logistics & Install", "Truck, rigging hardware, 3-person install crew, 4–6 hr install window. Lead technician takes the supervisor role at load-in.", 4800)
    row = line_item(ws, row, "On-Site Technicians", "Dedicated lead technician plus rotating second tech for rush windows and breaks, covering the 11 AM–8 PM operating window. Includes pre-open soundcheck and post-close shutdown.", 3000)
    row = line_item(ws, row, "Same-Day Strike & Outbound Freight", "Complete de-installation, module breakdown, crated outbound freight within the Lincoln Road BID's contracted strike window. Site walk-through with the BID operations team.", 3300)
    row = subtotal_row(ws, row, "Subtotal — The Activation (Lincoln Road Mall)", 11100)

    row = phase_header(ws, row, "Phase 04b  —  Logistics to NYC")
    row = line_item(ws, row, "Climate-Controlled Warehouse Hold", "Secure climate-controlled storage at AGV Miami's NY shop between Miami strike and Gymshark NYC flagship delivery.", 2400)
    row = line_item(ws, row, "Inter-City Freight (Climate-Controlled Truck)", "Miami → AGV Miami NY shop with real-time GPS tracking, two-driver rotation, handoff documentation.", 5500)
    row = line_item(ws, row, "Light Touchup at NYC Staging", "Light cosmetic touchup pass at AGV Miami's NY shop following Miami strike: scuff/scratch repair, paint colour-match, lightbox edge cleanup, IVR/camera/printer functional re-test.", 3000)
    row = line_item(ws, row, "White-Glove Delivery to Gymshark NYC Flagship", "Climate-controlled final-mile delivery from AGV Miami's NY shop to the Gymshark NYC flagship retail address. Scheduled overnight or pre-open per store operations.", 2000)
    row = subtotal_row(ws, row, "Subtotal — Logistics to NYC", 12900)

    row = phase_header(ws, row, "Phase 05  —  Project Management & Client Services")
    row = line_item(ws, row, "Project Management Fee", "Dedicated senior producer, weekly status reporting, milestone tracking, vendor / venue liaison, COI + insurance coordination, post-event reconciliation. Async-first cadence appropriate for a single-event programme.", 6000)
    row = subtotal_row(ws, row, "Subtotal — Project Management", 6000)

    row += 1
    row = gross_total_row(ws, row, "Gross Production Investment (Base)", 84550)
    row = credit_row(ws, row, "Less: Preferred Partner Credit", -5000)
    row = net_total_bar(ws, row, "NET V4.0 PRODUCTION INVESTMENT (BASE)", 79550)
    row += 1

    # Optional add-on block
    row = phase_header(ws, row, "Optional Add-On  —  NYC Local Install Team (Available, Not in Base)")
    row = line_item(ws, row, "Local Install Team — 1 Lead + 2 Crew", "On-site retail-environment install team for the Gymshark NYC flagship: 1 lead technician + 2 crew, floor-protection kit, retail-grade hand-tool kit, store-operations liaison. 4-hour install window scheduled to avoid trading hours.", 2800)
    row = line_item(ws, row, "In-Store Fixture Setup", "Anchoring to retail-spec floor (no permanent penetration), electrical drop, IVR + camera + printer + lightbox bring-up, on-floor commissioning sign-off with the flagship store manager.", 1600)
    row = subtotal_row(ws, row, "Subtotal — Optional Local Install Team", 4400)
    row += 1
    row = gross_total_row(ws, row, "Net V4.0 with Optional Local Install Team", 83950)
    row += 1

    row = payment_block(ws, row, deposit=47730, balance=31820, balance_due_date="on or before Friday, July 10, 2026", version_short="V4.0")
    footer(ws, row, "V4.0 — Lincoln Road Mall, Friday, July 17, 2026")


# ---------- V2.0 SHEET ----------
def build_v2_sheet(ws):
    setup(ws)
    ws.sheet_properties.tabColor = TAB_TEAL
    row = header_block(
        ws, row=1,
        version="Version 2.0",
        issue_date="April 29, 2026",
        scope_summary=(
            "Two-event sequence. Event 01 — Miami: Lincoln Road Mall street activation on Saturday, June 27, 2026 "
            ". Event 02 — New York: in-store fixture at Gymshark "
            "Bond St, delivered by July 9, 2026 to anchor the July 11 product launch and a 4-week in-store run "
            "through August 6, 2026. Inter-city respray/rewrap to NYC palette between events. "
            "[Superseded by V3.0 — included for reference only.]"
        ),
        activation_summary=("Event 01 (Miami): June 27, 2026", "Event 02 (NYC Bond St): July 9 – Aug 6, 2026"),
    )

    row = phase_header(ws, row, "Phase 01  —  The Phone Box")
    row = line_item(ws, row, "Phone Box Structural Shell", "Custom scenic fabrication, marine-grade paint, two-piece modular", 14500)
    row = line_item(ws, row, "4-Sided Illuminated Lightbox Signage", "LED-backlit translucent face panels on all four sides", 6800)
    row = line_item(ws, row, "Glass & Semi-Transparent Vinyl Panels", "Tempered glass on front door + sides", 2900)
    row = line_item(ws, row, "Pink Vinyl Interior Lining", "Four-wall pink vinyl wrap with 'bum mirror' messaging", 3200)
    row = line_item(ws, row, "Concealed Sliding Prize Door", "Motorised, colour-matched", 3400)
    row = line_item(ws, row, "Interior Finishes", "Chequer plate floor, dome light, seat, analogue phone", 2750)
    row = line_item(ws, row, "Engineering, Structural Calcs & Shop Drawings", "Wind-load, ballast, electrical, CAD", 4000)
    row = subtotal_row(ws, row, "Subtotal — The Phone Box", 37550)

    row = phase_header(ws, row, "Phase 02  —  Interactive Tech + Content Capture")
    row = line_item(ws, row, "Pre-Recorded Call-Response (IVR) System", "Multi-branch scripting, licensed voice, QA", 5500)
    row = line_item(ws, row, "Live Call / Walkie-Talkie Relay", "Dual-mode encrypted relay, backup uplink, producer headset", 2900)
    row = line_item(ws, row, "Ceiling Dual Camera System", "Two 4K cameras, cloud storage, live-preview, disclosure signage", 4500)
    row = line_item(ws, row, "Selfie & Belfie Mobile Mounts", "Motion-triggered ring-light inserts, dual-height", 1800)
    row = line_item(ws, row, "Thermal Voucher Printer & Shelf Mount", "Dual-speed thermal printer, pink wrap, redundant roll inventory", 2800)
    row = line_item(ws, row, "Content Capture & Media Handoff Platform", "Cloud workspace, proxies, 48-hour handoff", 2500)
    row = subtotal_row(ws, row, "Subtotal — Interactive Tech + Content Capture", 20000)

    row = phase_header(ws, row, "Phase 03  —  Branding, Signage & Consumables")
    row = line_item(ws, row, "Disclosure & Wayfinding Signage", "Camera-disclosure, queue management decals, brand lockup callouts", 950)
    row = line_item(ws, row, "Venue Dressing Kit", "A-frame signage, branded stanchions, pavement decals", 1750)
    row = line_item(ws, row, "Daily Consumables & Spares Kit", "Voucher rolls, cleaning, spares, gaffer", 2500)
    row = subtotal_row(ws, row, "Subtotal — Branding, Signage & Consumables", 5200)

    row = phase_header(ws, row, "Phase 04a  —  Event 01: Miami Lincoln Road Street Activation")
    row = line_item(ws, row, "Inbound Logistics & Install (Miami)", "3-person install crew, permit-window supervisor, 4–6 hr install window", 5800)
    row = line_item(ws, row, "On-Site Technicians (Miami)", "Lead technician + rotating second tech for rush windows", 2800)
    row = line_item(ws, row, "Same-Day Strike & Outbound Freight (Miami)", "De-installation, module breakdown, crated outbound freight", 3300)
    row = subtotal_row(ws, row, "Subtotal — Event 01 Miami", 11900)

    row = phase_header(ws, row, "Phase 04b  —  Inter-City Respray/Rewrap to NYC Palette")
    row = line_item(ws, row, "Climate-Controlled Warehouse Hold (~10 days)", "Secure climate-controlled storage between Miami strike and NYC respray", 2900)
    row = line_item(ws, row, "Inter-City Freight (Miami → NYC, Climate Truck)", "Dedicated climate truck with two-driver rotation, real-time GPS tracking, handoff documentation", 5500)
    row = line_item(ws, row, "Exterior Respray to NYC Palette", "Sand & prep, marine-grade respray to NYC product-launch colour callout, clear over-laminate", 5400)
    row = line_item(ws, row, "Interior Vinyl Refresh (NYC Product Launch Creative)", "Removal of Miami interior vinyl, application of NYC product-launch creative, mirror messaging update", 3200)
    row = line_item(ws, row, "Lightbox Graphic Refresh (NYC Launch Creative)", "Reprint and reinstall four translucent lightbox faces with NYC launch wordmark and treatment", 2400)
    row = line_item(ws, row, "Retail Compliance & Store Re-Spec", "Anchoring/ballast for indoor floor, electrical recheck, fire-marshal review, COI re-issuance for Bond St", 2000)
    row = line_item(ws, row, "Pre-Deployment QC & Retail-Grade Certification", "Full reassembly inspection, IVR + camera + printer bench re-test, retail-grade finish certification", 2600)
    row = subtotal_row(ws, row, "Subtotal — Respray Package", 24000)

    row = phase_header(ws, row, "Phase 04c  —  Event 02: Gymshark Bond St Delivery & 4-Week In-Store Run")
    row = line_item(ws, row, "White-Glove Delivery to Gymshark Bond St", "Climate-controlled delivery to 11 Bond St, NYC, scheduled overnight or pre-open. Target on or before July 9, 2026.", 3000)
    row = line_item(ws, row, "Retail-Environment Install Crew", "2-person crew with floor-protection, retail-grade hand-tool kit, store operations liaison, 4-hour install window", 2300)
    row = line_item(ws, row, "In-Store Fixture Setup", "Retail-spec floor anchoring, electrical drop, IVR + camera + printer + lightbox bring-up, on-floor sign-off with store manager", 2100)
    row = line_item(ws, row, "Store Wayfinding & Footfall Driver Kit", "Sidewalk A-frame, branded window-vinyl tease, in-store directional decals, printed/social CTA pack", 2200)
    row = line_item(ws, row, "4-Week In-Store Presence Support", "Remote tech monitoring, weekly voucher-roll restock, content uploads, scenic touch-ups through August 6, 2026", 2500)
    row = line_item(ws, row, "Retail Strike & Return Freight", "End-of-launch de-installation, retail-grade floor restoration, return freight to Bohemia, NY facility", 1400)
    row = subtotal_row(ws, row, "Subtotal — Event 02 Bond St", 13500)

    row = phase_header(ws, row, "Phase 05  —  Project Management & Client Services")
    row = line_item(ws, row, "Project Management Fee", "Dedicated senior producer, weekly status, milestone tracking, COI + insurance coordination", 12000)
    row = subtotal_row(ws, row, "Subtotal — Project Management", 12000)

    row += 1
    row = gross_total_row(ws, row, "Gross Production Investment", 124150)
    row = credit_row(ws, row, "Less: Preferred Partner Credit", -5000)
    row = net_total_bar(ws, row, "NET V2.0 PRODUCTION INVESTMENT", 119150)
    row += 1

    row = payment_block(ws, row, deposit=71490, balance=47660, balance_due_date="five (5) business days prior to Event 01 install", version_short="V2.0")
    footer(ws, row, "V2.0 — superseded by V3.0")


# ---------- V1.0 SHEET ----------
def build_v1_sheet(ws):
    setup(ws)
    ws.sheet_properties.tabColor = TAB_MUTED
    row = header_block(
        ws, row=1,
        version="Version 1.0",
        issue_date="April 17, 2026",
        scope_summary=(
            "Initial issuance: three execution options against a $120,000 overall budget. Option 01 NYC-only; "
            "Option 02 Miami-only; Option 03 dual-city (Miami first → NYC second) with full inter-city reskin "
            "package. Twelve recommended venues across both cities. The figures below itemize the recommended "
            "Option 03 dual-city configuration ($98,500 production target). [Superseded by V2.0 and V3.0 — "
            "included for reference only.]"
        ),
        activation_summary=("Recommended Option 03: Miami Jun 27, 2026", "→ NYC Jul 11, 2026 (with reskin)"),
    )

    row = phase_header(ws, row, "Phase 01  —  The Phone Box")
    row = line_item(ws, row, "Phone Box Structural Shell", "Custom scenic fabrication, marine-grade paint, two-piece modular", 14500)
    row = line_item(ws, row, "4-Sided Illuminated Lightbox Signage", "LED-backlit translucent face panels on all four sides", 6800)
    row = line_item(ws, row, "Glass & Semi-Transparent Vinyl Panels", "Tempered glass on front door + sides", 2900)
    row = line_item(ws, row, "Pink Vinyl Interior Lining", "Four-wall pink vinyl wrap with 'bum mirror' messaging", 3200)
    row = line_item(ws, row, "Concealed Sliding Prize Door", "Motorised, colour-matched", 3400)
    row = line_item(ws, row, "Interior Finishes", "Chequer plate floor, dome light, seat, analogue phone", 2750)
    row = subtotal_row(ws, row, "Subtotal — The Phone Box", 33550)

    row = phase_header(ws, row, "Phase 02  —  Interactive Tech")
    row = line_item(ws, row, "Pre-Recorded Call-Response (IVR) System", "Scripting, voice, keypad logic", 3500)
    row = line_item(ws, row, "Live Call / Walkie-Talkie Relay", "Social producer handoff", 1800)
    row = line_item(ws, row, "Ceiling Dual Camera System", "Video+audio, cloud storage, disclosure sign", 2400)
    row = line_item(ws, row, "Thermal Voucher Printer & Shelf Mount", "Wrapped, shelf aligned to sliding door", 1950)
    row = line_item(ws, row, "Selfie & Belfie Mobile Mounts", "Colour-matched, dual-height", 650)
    row = subtotal_row(ws, row, "Subtotal — Interactive Tech", 10300)

    row = phase_header(ws, row, "Phase 03  —  Branding & Packaging")
    row = line_item(ws, row, "Custom Product Boxes (200 units)", "Yellow-pages inspired, 300×200×70 mm", 4800)
    row = line_item(ws, row, "Disclosure & Wayfinding Signage", "Camera disclosure, queue management decals", 950)
    row = line_item(ws, row, "Venue Dressing Kit", "A-frames, stanchions, pavement decals", 1750)
    row = subtotal_row(ws, row, "Subtotal — Branding & Packaging", 7500)

    row = phase_header(ws, row, "Phase 04  —  Logistics, Install & Strike (Per Activation)")
    row = line_item(ws, row, "Inbound Logistics & Install — Miami", "Truck, rigging, install crew, electrical hookup", 4200)
    row = line_item(ws, row, "On-Site Technician — Miami", "Single technician across operating window", 1400)
    row = line_item(ws, row, "Same-Day Strike & Outbound Freight — Miami", "De-installation, module breakdown, freight", 2800)
    row = line_item(ws, row, "Inbound Logistics & Install — NYC", "Truck, rigging, install crew, electrical hookup", 4200)
    row = line_item(ws, row, "On-Site Technician — NYC", "Single technician across operating window", 1400)
    row = line_item(ws, row, "Same-Day Strike & Outbound Freight — NYC", "De-installation, module breakdown, freight", 2800)
    row = subtotal_row(ws, row, "Subtotal — Two-City Logistics", 16800)

    row = phase_header(ws, row, "Phase 04 Add-On  —  Inter-City Rewrap (Option 03)")
    row = line_item(ws, row, "Inter-City Rewrap & Transport", "Warehouse hold, vinyl refresh, Miami↔NYC truck", 9500)
    row = subtotal_row(ws, row, "Subtotal — Inter-City Rewrap", 9500)

    row = phase_header(ws, row, "V1.0 Note  —  Other Programme Elements (Bundled)")
    row = line_item(
        ws, row,
        "Programme Bundle (V1.0 reference)",
        "V1.0 published a $98,500 dual-city production target that included additional logistics, contingency, and venue-specific finishing not yet itemized at issuance. The line items above ($77,650) plus this bundled allowance ($20,850) reconcile to the published $98,500 V1.0 figure. Subsequent versions broke out these items individually (engineering, content capture, project management, retail/respray scope changes).",
        20850,
    )
    row = subtotal_row(ws, row, "Subtotal — Bundled Allowance", 20850)

    row += 1
    row = gross_total_row(ws, row, "V1.0 Production Investment (Option 03 Dual-City)", 98500)
    row += 1

    row = payment_block(ws, row, deposit=59100, balance=39400, balance_due_date="three (3) business days prior to first activation", version_short="V1.0")
    footer(ws, row, "V1.0 — superseded by V2.0 and V3.0")


def main():
    # 1. Single-tab V3.0 invoice (current scope only)
    wb_v3 = Workbook()
    ws3 = wb_v3.active
    ws3.title = "V3.0 — Issued"
    build_v3_sheet(ws3)
    wb_v3.save(OUTPUT_V3)
    print(f"Wrote {OUTPUT_V3} with sheets: {[s.title for s in wb_v3.worksheets]}")

    # 2. Single-tab V3.1 invoice (preceding proposed update)
    wb_v31 = Workbook()
    ws31 = wb_v31.active
    ws31.title = "V3.1 — Proposed"
    build_v31_sheet(ws31)
    wb_v31.save(OUTPUT_V31)
    print(f"Wrote {OUTPUT_V31} with sheets: {[s.title for s in wb_v31.worksheets]}")

    # 3. V4.0 invoice — multi-tab (current proposed + V3.0 issued + V2.0 superseded for reference)
    wb_v40 = Workbook()
    ws40 = wb_v40.active
    ws40.title = "V4.0 — Current Proposed"
    build_v40_sheet(ws40)

    ws40_v3 = wb_v40.create_sheet(title="V3.0 — Issued")
    build_v3_sheet(ws40_v3)

    ws40_v2 = wb_v40.create_sheet(title="V2.0 — Superseded")
    build_v2_sheet(ws40_v2)

    wb_v40.save(OUTPUT_V40)
    print(f"Wrote {OUTPUT_V40} with sheets: {[s.title for s in wb_v40.worksheets]}")

    # 4. Five-tab version history workbook (V4.0 → V3.1 → V3.0 → V2.0 → V1.0)
    wb_history = Workbook()

    ws40h = wb_history.active
    ws40h.title = "V4.0 — Value-Engineered"
    build_v40_sheet(ws40h)

    ws31h = wb_history.create_sheet(title="V3.1 — Proposed")
    build_v31_sheet(ws31h)

    ws3h = wb_history.create_sheet(title="V3.0 — Issued")
    build_v3_sheet(ws3h)

    ws2h = wb_history.create_sheet(title="V2.0 — Superseded")
    build_v2_sheet(ws2h)

    ws1h = wb_history.create_sheet(title="V1.0 — Initial")
    build_v1_sheet(ws1h)

    wb_history.save(OUTPUT_HISTORY)
    print(f"Wrote {OUTPUT_HISTORY} with sheets: {[s.title for s in wb_history.worksheets]}")


if __name__ == "__main__":
    main()
